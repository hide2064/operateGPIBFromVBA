"""
create_excel.py -- GPIB制御用Excelファイル生成スクリプト

使い方:
    python create_excel.py
    python create_excel.py --output MyGpib.xlsx

生成されるファイル:
    Config シート  : 機器設定 (GPIB/LAN サンプルデータ入り)
    Control シート : コマンド実行画面 (サンプルコマンド入り)
    Setup シート   : VBAインポート手順・ボタン設定方法

出力後の手順:
    1. 生成された .xlsx を Excel で開く
    2. 名前を付けて保存 -> "Excel マクロ有効ブック (*.xlsm)" で保存し直す
    3. VBAエディタ (Alt+F11) でモジュールをインポート
    4. Setup シートの手順に従ってボタンを作成する
"""
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ colors --
C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_SAMPLE_BG = "F2F7FB"
C_BORDER    = "BFBFBF"


def _side():
    return Side(style="thin", color=C_BORDER)


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _hfont(size=10):
    return Font(name="Meiryo UI", bold=True, color=C_HEADER_FG, size=size)


def _bfont(bold=False, color="000000", size=10):
    return Font(name="Meiryo UI", bold=bold, color=color, size=size)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _center():
    return Alignment(horizontal="center", vertical="center")


def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _header_row(ws, row, cols: dict):
    for col, label in cols.items():
        c = ws.cell(row=row, column=col, value=label)
        c.font      = _hfont()
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()


def _data_row(ws, row, data: dict, bg=None):
    for col, value in data.items():
        c = ws.cell(row=row, column=col, value=value)
        c.font      = _bfont()
        c.alignment = _left()
        c.border    = _border()
        if bg:
            c.fill = _fill(bg)


# --------------------------------------------------------------- Config -----
def build_config_sheet(wb: Workbook):
    ws = wb.create_sheet("Config")
    ws.sheet_view.showGridLines = False

    for col, w in {1: 20, 2: 32, 3: 12, 4: 14, 5: 20, 6: 10}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    t = ws.cell(row=1, column=1, value="Config -- 機器設定")
    t.font      = Font(name="Meiryo UI", bold=True, size=13, color=C_HEADER_BG)
    t.alignment = _left()
    ws.merge_cells("A1:F1")

    n = ws.cell(row=2, column=1,
                value="B列(VISAアドレス)はフル指定。D/E列を入力した場合は自動生成されB列より優先されます。")
    n.font      = _bfont(size=9, color="595959")
    n.alignment = _left()
    ws.merge_cells("A2:F2")

    ws.row_dimensions[3].height = 20
    _header_row(ws, 3, {
        1: "機器名 (Name)",
        2: "VISAアドレス (B列フル指定)",
        3: "Timeout(ms)",
        4: "Protocol (D列)",
        5: "Host / IP (E列)",
        6: "Port (F列)",
    })

    samples = [
        {1: "DMM",             2: "GPIB0::22::INSTR", 3: 5000,  4: "",       5: "",             6: ""},
        {1: "PowerSupply",     2: "GPIB0::5::INSTR",  3: 3000,  4: "",       5: "",             6: ""},
        {1: "Scope",           2: "",                 3: 10000, 4: "TCPIP",  5: "192.168.1.10", 6: ""},
        {1: "MT8821C",         2: "",                 3: 10000, 4: "SOCKET", 5: "192.168.1.20", 6: "5025"},
        {1: "SpectrumAnalyzer",2: "",                 3: 10000, 4: "HISLIP", 5: "192.168.1.30", 6: ""},
    ]
    for i, row_data in enumerate(samples):
        r = i + 4
        ws.row_dimensions[r].height = 18
        _data_row(ws, r, row_data, bg=C_SAMPLE_BG if i % 2 == 0 else None)

    r = len(samples) + 4
    for col in range(1, 7):
        c = ws.cell(row=r, column=col, value="<-- ここに追加")
        c.font      = _bfont(color="BFBFBF", size=9)
        c.alignment = _left()
        c.border    = _border()


# -------------------------------------------------------------- Control -----
def build_control_sheet(wb: Workbook):
    ws = wb.create_sheet("Control")
    ws.sheet_view.showGridLines = False

    for col, w in {1: 20, 2: 30, 3: 40, 4: 20}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    t = ws.cell(row=1, column=1, value="Control -- コマンド実行")
    t.font      = Font(name="Meiryo UI", bold=True, size=13, color=C_HEADER_BG)
    t.alignment = _left()
    ws.merge_cells("A1:D1")

    n = ws.cell(row=2, column=1,
                value="A列=Config機器名, B列=SCPIコマンド ('?' 末尾でクエリ), C列=応答, D列=ステータス")
    n.font      = _bfont(size=9, color="595959")
    n.alignment = _left()
    ws.merge_cells("A2:D2")

    b = ws.cell(row=3, column=1,
                value="▼ この行付近にボタンを挿入してマクロを割り当ててください (Setup シート参照)")
    b.font      = Font(name="Meiryo UI", bold=True, color="FF6600", size=9)
    b.fill      = _fill("FFF2CC")
    b.alignment = _left()
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:D3")

    ws.row_dimensions[4].height = 20
    _header_row(ws, 4, {1: "機器名", 2: "SCPIコマンド", 3: "応答結果", 4: "ステータス"})

    samples = [
        {1: "DMM",              2: "*IDN?",         3: "", 4: ""},
        {1: "DMM",              2: "*RST",           3: "", 4: ""},
        {1: "DMM",              2: "MEAS:VOLT:DC?",  3: "", 4: ""},
        {1: "PowerSupply",      2: "*IDN?",         3: "", 4: ""},
        {1: "PowerSupply",      2: "VOLT 5.0",      3: "", 4: ""},
        {1: "PowerSupply",      2: "OUTP ON",       3: "", 4: ""},
        {1: "Scope",            2: "*IDN?",         3: "", 4: ""},
        {1: "MT8821C",          2: "*IDN?",         3: "", 4: ""},
        {1: "MT8821C",          2: "*RST",          3: "", 4: ""},
        {1: "SpectrumAnalyzer", 2: "*IDN?",         3: "", 4: ""},
    ]
    for i, row_data in enumerate(samples):
        r = i + 5
        ws.row_dimensions[r].height = 18
        _data_row(ws, r, row_data, bg=C_SAMPLE_BG if i % 2 == 0 else None)


# --------------------------------------------------------------- Result -----
def build_result_sheet(wb: Workbook):
    ws = wb.create_sheet("Result")
    ws.sheet_view.showGridLines = False

    col_widths = {1: 6, 2: 20, 3: 16, 4: 32, 5: 14, 6: 24, 7: 36, 8: 10, 9: 20}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    _header_row(ws, 1, {
        1: "No.",
        2: "実行日時",
        3: "機器名",
        4: "VISAアドレス",
        5: "接続方式",
        6: "コマンド / アクション",
        7: "応答結果",
        8: "ステータス",
        9: "備考",
    })

    # サンプル行 (イメージ表示用)
    sample_rows = [
        {1: 1, 2: "2025/01/01 12:00:00", 3: "MT8821C", 4: "GPIB0::1::INSTR",
         5: "GPIB", 6: "*IDN?", 7: "ANRITSU,MT8821C,...", 8: "OK", 9: ""},
        {1: 2, 2: "2025/01/01 12:00:01", 3: "MT8821C", 4: "TCPIP0::192.168.1.10::INSTR",
         5: "LAN VXI-11", 6: "BAND 1", 7: "", 8: "OK", 9: ""},
        {1: 3, 2: "2025/01/01 12:00:02", 3: "MT8821C", 4: "TCPIP0::192.168.1.10::INSTR",
         5: "LAN VXI-11", 6: "BAND?", 7: "1", 8: "OK", 9: ""},
    ]
    for i, row_data in enumerate(sample_rows):
        r = i + 2
        ws.row_dimensions[r].height = 18
        _data_row(ws, r, row_data, bg=C_SAMPLE_BG if i % 2 == 0 else None)
        # ステータス列を緑色に
        ws.cell(row=r, column=8).font = Font(name="Meiryo UI", bold=True,
                                             color="006400", size=10)

    # 注記
    note_r = len(sample_rows) + 2
    note = ws.cell(row=note_r, column=1,
                   value="(上記はサンプル表示です。実行時には自動的に追記されます。"
                         " ResultSheet.bas をインポートして使用してください。)")
    note.font      = Font(name="Meiryo UI", color="BFBFBF", italic=True, size=9)
    note.alignment = _left()
    ws.merge_cells(f"A{note_r}:I{note_r}")


# ---------------------------------------------------------------- Setup -----
def build_setup_sheet(wb: Workbook, vba_dir: str):
    ws = wb.create_sheet("Setup")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    def t(row, text, fg=C_HEADER_BG, bg=None, size=13):
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Meiryo UI", bold=True, size=size, color=fg)
        c.alignment = _left()
        if bg:
            c.fill = _fill(bg)
        ws.row_dimensions[row].height = 22

    def b(row, text, bold=False, color="000000", bg=None, size=10):
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Meiryo UI", bold=bold, size=size, color=color)
        c.alignment = _left(wrap=True)
        if bg:
            c.fill = _fill(bg)
        ws.row_dimensions[row].height = 16

    r = 1
    t(r, "セットアップ手順", size=14); r += 1
    b(r, "このシートを参考に VBA モジュールのインポートとボタン設定を行ってください。",
      size=9, color="595959"); r += 2

    # STEP 1
    t(r, "STEP 1: .xlsm 形式で保存し直す", bg=C_HEADER_BG, fg="FFFFFF", size=11); r += 1
    b(r, "1. ファイル -> 名前を付けて保存"); r += 1
    b(r, '2. ファイルの種類を「Excel マクロ有効ブック (*.xlsm)」に変更して保存'); r += 1
    b(r, "3. 保存後、このファイルを再度開く"); r += 2

    # STEP 2
    t(r, "STEP 2: VBA モジュールをインポートする", bg=C_HEADER_BG, fg="FFFFFF", size=11); r += 1
    b(r, "1. Alt + F11 -> VBAエディタを開く"); r += 1
    b(r, "2. 「ファイル」->「ファイルのインポート」で下記を順番にインポートする:"); r += 1

    bas_files = [
        ("AppConfig.bas",       "設定ファイル読み込み (必須)"),
        ("GpibControlHttp.bas", "Flask方式 実行モジュール (推奨)"),
        ("GpibControl.bas",     "CLI方式 実行モジュール (試験用・任意)"),
        ("GpibMT8821C.bas",     "Anritsu MT8821C 専用モジュール (任意)"),
        ("ResultSheet.bas",     "試験結果 Result シート管理モジュール (任意)"),
        ("MT8821C_Sample.bas",  "MT8821C 動作確認サンプル集 (任意)"),
    ]
    for fname, desc in bas_files:
        exists = "OK" if os.path.exists(os.path.join(vba_dir, fname)) else "?"
        b(r, f"    [{exists}]  {fname}  --  {desc}", bg="F0F4F9"); r += 1

    b(r, f"VBAファイルの場所: {vba_dir}"); r += 2

    # STEP 3
    t(r, "STEP 3: settings.ini のパスを確認する", bg=C_HEADER_BG, fg="FFFFFF", size=11); r += 1
    b(r, "AppConfig.bas は「Excelファイルと同じフォルダ\\config\\settings.ini」を自動で読み込みます。"); r += 1
    ini_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config", "settings.ini")
    b(r, f"settings.ini の場所: {ini_path}", bg="F0F4F9"); r += 1
    b(r, "Excelファイルをリポジトリのルート (README.md と同じフォルダ) に保存すれば自動で見つかります。"); r += 2

    # STEP 4
    t(r, "STEP 4: ボタンを作成してマクロを割り当てる", bg=C_HEADER_BG, fg="FFFFFF", size=11); r += 1
    b(r, "Control シートで「挿入」->「図形」-> 長方形を描き、右クリック ->「マクロの登録」で割り当てる:"); r += 1

    btns = [
        ("サーバー起動",           "GpibControlHttp.StartGpibServer"),
        ("選択行を実行",           "GpibControlHttp.ExecuteSelectedCommandHttp"),
        ("すべて実行",             "GpibControlHttp.ExecuteAllCommandsHttp"),
        ("---", "---"),
        ("Control -> Result 転記", "ResultSheet.Result_AppendFromControl"),
        ("Result クリア",          "ResultSheet.Result_Clear"),
    ]
    for btn_name, macro in btns:
        if btn_name == "---":
            b(r, "    (Result シート操作ボタンを Result シートに追加する場合:)"); r += 1
        else:
            b(r, f"    ボタン名: 「{btn_name}」  ->  マクロ: {macro}", bg="F0F4F9"); r += 1
    r += 1

    # STEP 5
    t(r, "STEP 5: 動作確認", bg=C_HEADER_BG, fg="FFFFFF", size=11); r += 1
    b(r, "1. start_server.bat を実行して Flask サーバーを起動する"); r += 1
    b(r, "2. Config シートに実際の機器名とアドレスを入力する"); r += 1
    b(r, "3. Control シートのコマンド行を選択して「選択行を実行」ボタンを押す"); r += 1
    b(r, "4. D列に「OK」と表示されれば成功"); r += 1
    b(r, "5. 「Control -> Result 転記」ボタンで Result シートへ記録する"); r += 2

    # デバッグ
    t(r, "デバッグのヒント", size=12); r += 1
    b(r, "ログ確認: logs\\gpib.log  (PowerShell: Get-Content logs\\gpib.log -Wait -Tail 30)"); r += 1
    b(r, "API直接テスト: python debug_client.py health / execute / connections / log"); r += 1
    b(r, "サーバー状態確認: ブラウザで http://127.0.0.1:5000/debug を開く"); r += 1


# ------------------------------------------------------------------ main ----
def main():
    parser = argparse.ArgumentParser(description="GPIB制御用Excelファイルを生成する")
    parser.add_argument("--output", default="GpibControl.xlsx",
                        help="出力ファイル名 (デフォルト: GpibControl.xlsx)")
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    vba_dir  = os.path.join(base_dir, "vba")

    print("Config シートを生成中...")
    build_config_sheet(wb)

    print("Control シートを生成中...")
    build_control_sheet(wb)

    print("Result シートを生成中...")
    build_result_sheet(wb)

    print("Setup シートを生成中...")
    build_setup_sheet(wb, vba_dir)

    out_path = os.path.join(base_dir, args.output)
    wb.save(out_path)
    print(f"\n生成完了: {out_path}")
    print("\n【次のステップ】")
    print("  1. 上記ファイルを Excel で開く")
    print('  2. 「名前を付けて保存」-> "Excel マクロ有効ブック (*.xlsm)" で保存')
    print("  3. Alt+F11 で VBAエディタを開き、Setup シートの手順に従う")
    print("  4. Config シートに実機器のアドレスを入力して動作確認")


if __name__ == "__main__":
    main()
